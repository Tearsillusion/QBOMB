import tkinter
from tkinter import filedialog

import win32api
import win32ui
from PyQt5.QtWidgets import QWidget, QMessageBox, QButtonGroup
from send.send import Ui_qq_send
from send_success.send_success_init import QqSendSuccess
import time
from threading import Thread
import win32gui
import win32con
import win32clipboard as w
import openpyxl
from operation_configs import OperationConfigs

class QqSend(Ui_qq_send,QWidget):
    def __init__(self):
        super(QqSend, self).__init__()
        self.setupUi(self)
        self.pushButton_2.clicked.connect(self.close)
        self.send_submit.clicked.connect(self.startThread)
        self.thread_all_window = QqSendSuccess()
        self.thread_all_window.pushButton_2.clicked.connect(self.closeMain)
        self.send_Flage = True
        self.operation_config = OperationConfigs()
        # 类型
        self.button_group = QButtonGroup(self)
        self.button_group.addButton(self.radioButton,0)
        self.button_group.addButton(self.radioButton_2, 1)
        self.button_group.buttonClicked[int].connect(self.buttonGroupMain)
        self.send_type = 0
        # 获取excle文件
        self.excel_send.clicked.connect(self.excelSend)
        # excle模板下载
        self.excel_send_model.clicked.connect(self.excelDownload)
    # excel下载
    def excelDownload(self):
        tk = tkinter.Tk()
        tk.withdraw()
        path = filedialog.askdirectory()
        wb = openpyxl.Workbook()
        # 获取活跃的工作表，ws代表wb(工作簿)的一个工作表
        ws = wb.active
        # 更改工作表ws的title
        ws.title = 'test_sheet1'
        # 对ws的单个单元格传入数据
        data = ['信息1','信息2','信息3']
        data_excel = []
        # 将字典中的每对数据（键，值）以列表形式传入data_excel列表
        for each in data:
            data_excel.append([each])
        # 将data_excel列表内的内容存入工作表
        for each in data_excel:
            ws.append(each)
        # 注意：上述两个append方法是意义完全不同的两个方法
        wb.save(path+'/QBOMB@Excel发送模板.xlsx')
        QMessageBox().information(None, "提示", "下载成功！", QMessageBox.Yes)
    # 选择excel
    def excelSend(self):
        if self.lineEdit_3.text() == "":
            QMessageBox.warning(self, "警告", "请输入昵称", QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
        elif int(self.spinBox_2.text()) <= 0:
            QMessageBox.warning(self, "警告", "时间间隔需大于0", QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
        else:
            dlg = win32ui.CreateFileDialog(1, None, None)  # 指定为打开文件窗口
            dlg.SetOFNInitialDir("C:")
            dlg.DoModal()
            path = dlg.GetPathName()
            if path.index('xlsx') > -1:
                self.excelGetInfo(path)
            else:
                QMessageBox.warning(self, "警告", "你选择的不是Excel表格", QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
    # 获取excel内容并发送
    def excelGetInfo(self, path):

        readbook = openpyxl.load_workbook(path)#读取文件
        readbook_names = readbook.get_sheet_names()#得到excel有几个sheet内容
        print(readbook_names)
        self.content = readbook.get_sheet_by_name(readbook_names[0])#获取excel第一个表格
        max_row = self.content.max_row#获取内容列数长度
        print(max_row,self.content.cell(1, 1).value)
        # 发送信息
        self.sendMsg('Excel', max_row)
    # 获取发送类型
    def buttonGroupMain(self,id):
        self.send_type = id
    # 关闭发送
    def closeMain(self):
        self.send_Flage = False
        self.show()

    # 发送信息
    def startThread(self):
        if self.lineEdit_3.text() == "":
            QMessageBox.warning(self, "警告", "请输入昵称", QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
        elif int(self.spinBox_2.text()) <= 0:
            QMessageBox.warning(self, "警告", "时间间隔需大于0", QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
        elif int(self.spinBox.text()) <= 0:
            QMessageBox.warning(self, "警告", "次数需大于0", QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
        elif self.textEdit.toPlainText() == "":
            QMessageBox.warning(self, "警告", "请输入发送内容", QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
        else:
            self.send_Flage = True
            self.close()
            self.sendMsg()

    def sendMsg(self,type,max_row):

            self.receiver = self.lineEdit_3.text() #昵称
            self.msg = self.textEdit.toPlainText() #内容
            self.send_space = int(self.spinBox_2.text()) #间隔
            self.send_num = int(self.spinBox.text()) #次数
            self.thread_all_window.show()
            self.thread_all_window.label_6.setText(self.receiver)
            self.thread_all_window_list = self.thread_all_window
            if type=="Excel":
                self._thread = Thread(target=self.__run(type,max_row))
                self._thread.start()  # 启动线程
            else:
                self.setText()
                self._thread = Thread(target=self.__run)
                self._thread.start()  # 启动线程

    def __run(self,type,max_row):
        print(type,max_row)
        if type == "Excel":
            for k in range(max_row):
                if self.send_Flage:
                    self.msg = self.content.cell(k+1, 1).value
                    time.sleep(self.send_space)
                    if self.send_type==0:
                        self.setText()
                        self.sendQqmsg()
                    else:
                        self.sendWeixinmsg()
                    self.thread_all_window_list.label_4.setText(str(k + 1) + '次')
                    self.thread_all_window_list.label_5.setText(str(max_row - (k + 1)) + '次')
                    self.thread_all_window_list.label_7.setText(str((max_row - (k + 1)) / self.send_space) + 's')
        else:
            for k in range(self.send_num):
                if self.send_Flage:
                    time.sleep(self.send_space)
                    if self.send_type==0:
                        self.sendQqmsg()
                    else:
                        self.sendWeixinmsg()
                    self.thread_all_window_list.label_4.setText(str(k + 1) + '次')
                    self.thread_all_window_list.label_5.setText(str(self.send_num - (k + 1)) + '次')
                    self.thread_all_window_list.label_7.setText(str((self.send_num - (k + 1)) / self.send_space) + 's')

                else:
                    break

    # 设置剪贴版内容
    def setText(self):
        w.OpenClipboard()
        w.EmptyClipboard()
        w.SetClipboardData(win32con.CF_UNICODETEXT, self.msg)
        w.CloseClipboard()

    def ctrlV(self):
        win32api.keybd_event(17, 0, 0, 0)  # ctrl键位码是17
        win32api.keybd_event(86, 0, 0, 0)  # v键位码是86
        win32api.keybd_event(86, 0, win32con.KEYEVENTF_KEYUP, 0)  # 释放按键
        win32api.keybd_event(17, 0, win32con.KEYEVENTF_KEYUP, 0)

    def enter(self):
        win32api.keybd_event(13, 0, 0, 0)  # Enter键位码
        win32api.keybd_event(13, 0, win32con.KEYEVENTF_KEYUP, 0)  # 释放按键
    # 发送QQ信息
    def sendQqmsg(self):
        qq = win32gui.FindWindow(None, self.receiver)
        if qq == 0:
            self.thread_all_window.hide()
            self.show()
            QMessageBox.warning(self, "警告", "昵称不存在/QQ好友对话框未打开", QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
            return
        win32gui.SetForegroundWindow(qq)
        win32gui.SendMessage(qq, win32con.WM_PASTE, 0, 0)
        win32gui.SendMessage(qq, win32con.WM_KEYDOWN, win32con.VK_RETURN, 0)  # 模拟按下Enter键
        win32gui.SendMessage(qq, win32con.WM_KEYUP, win32con.VK_RETURN, 0)  # 模拟松开
    # 发送微信消息
    def sendWeixinmsg(self):
        qq = win32gui.FindWindow(None, self.receiver)
        if qq == 0:
            self.thread_all_window.hide()
            self.show()
            QMessageBox.warning(self, "警告", "昵称不存在/QQ好友对话框未打开", QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
            return
        win32gui.SetForegroundWindow(qq)
        win32gui.ShowWindow(qq, win32con.WM_SHOWWINDOW)
        self.setText()
        time.sleep(1)
        self.ctrlV()
        self.enter()
        win32gui.SendMessage(qq, win32con.WM_KEYDOWN, win32con.VK_RETURN, 0)  # 模拟按下Enter键
        win32gui.SendMessage(qq, win32con.WM_KEYUP, win32con.VK_RETURN, 0)  # 模拟松开
